#This program will take in the Excel spreadsheet that ITS uses and output a formatted spreadsheet that can then be imported into Humanity as is.
from openpyxl import Workbook
from openpyxl import load_workbook
import AuxiliaryFunctions
import re
from datetime import datetime, timedelta

#TODO: Be able to toggle between Library shifts or helpdesk shifts only or both.
#load in workbook
filename = input("Place ITS Schedule Spreadsheet in the same folder as this program, then type its filename here (e.g Schedule.xlsx): ")
wb = load_workbook(filename)
#Note: also load in error handling?
scheduleSheet = wb.get_sheet_by_name('CarlTech Schedule')
staffSheet = wb.get_sheet_by_name('CarlTech Staff List')


#Get list of all the workers
workerList = []
staffNumber = staffSheet.max_row #subtract 1 to get actual number of workers, because the first row will always be the column headers for titles and whatnot.
for row in range(2, staffNumber): #for every worker
	firstName = staffSheet.cell(row = row, column = 1).value
	lastName = staffSheet.cell(row = row, column = 2).value
	workerList.append((firstName,lastName))


#create the output workbook
outputWb = Workbook()
outputSheet = outputWb.active
outputWb.save('HumanityScheduleImport.xlsx') #note: this will overwrite existing files without warning.

#populate column headers
outputSheet['A1'] = 'name'
outputSheet['B1'] = 'position'

startDate = AuxiliaryFunctions.dateParser(input("When does the term start? (mm/dd/yyyy): ")) #

for col in range(3,10):
	dateString = startDate.strftime('%m/%d/%y') #get start date as a string
	outputSheet.cell(row = 1, column = col, value =  dateString)
	startDate = startDate + timedelta(days=1)


# outputSheet['C1'] = '1/1/18' #TODO: automate populating dates based on eventual incorporation of user input
# outputSheet['D1'] = '1/2/18'
# outputSheet['E1'] = '1/3/18'
# outputSheet['F1'] = '1/4/18'
# outputSheet['G1'] = '1/5/18'
# outputSheet['H1'] = '1/6/18'
# outputSheet['I1'] = '1/7/18'


#for col in scheduleSheet['A:N']:
for col in range(1,15): #this corresponds to columns A through N
	#for curRow in scheduleSheet[2:52]:
	for curRow in range(2,53):
		workerName = scheduleSheet.cell(row = curRow, column = col).value
		timeRegex = re.compile('[^a-zA-z.\s]') #will cause any cell that doesn't contain a name to be skipped.
		if workerName != None:
			workerName = str(workerName) #just to make sure you're dealing with a string....
			if timeRegex.search(workerName) == None: #check that you're not actually dealing with a blank cell or a Time Cell
				name = AuxiliaryFunctions.getStudent(workerList, workerName)
				position = AuxiliaryFunctions.getPosition(col)
				day = AuxiliaryFunctions.getDay(col)
				shiftTime = AuxiliaryFunctions.getShiftTime(col, curRow)
				outputSheet.append((name, position))
				row_count = outputSheet.max_row
				if shiftTime == "12:00am-1:00am": #For those cases where the time actually stretches into the next day, we increment forward by one accordingly, unless it's at the end of the week (a sunday), in which case we set the day back to Monday.
					day += 1
					if day > 9:
						day = 3
				outputSheet.cell(row = row_count, column = day, value = shiftTime)

#if you want to expand the thing out to the end of term.
endDate = AuxiliaryFunctions.dateParser(input("When does the term end? (mm/dd/yyyy): ")) #
col = 9
lastEnteredDate = datetime.strptime(outputSheet.cell(row = 1, column = col).value, '%m/%d/%y').date()
maxRow = outputSheet.max_row
#populates column headers properly

while endDate != lastEnteredDate:
	col+=1
	lastEnteredDate = lastEnteredDate + timedelta(days=1)
	dateString = lastEnteredDate.strftime('%m/%d/%y') #get lastEnteredDate as a string
	outputSheet.cell(row = 1, column = col, value = dateString) 
	
	#copy the original scheduled column over.
	curRow = 2
	while curRow <= maxRow:
		originalCell = outputSheet.cell(row = curRow, column = col-7)
		newCell = outputSheet.cell(row = curRow, column = col)
		newCell.value = originalCell.value
		curRow += 1
	


print ("output saved to HumanityScheduleImport.xlsx")
outputWb.save('HumanityScheduleImport.xlsx')